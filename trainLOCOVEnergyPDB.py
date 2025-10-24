import tensorflow as tf
import numpy as np
from collections import defaultdict
import gc
from model import get_model
import os
from sklearn.model_selection import LeaveOneOut
import pandas as pd
import sys
from sklearn.model_selection import ShuffleSplit, StratifiedKFold, train_test_split
from sklearn.model_selection import StratifiedShuffleSplit, KFold
from sklearn import metrics
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from scipy.stats import pearsonr, spearmanr
from sklearn.linear_model import LinearRegression
import openpyxl as op
import matplotlib.pyplot as plt
import warnings
from tensorflow import keras
from sklearn.preprocessing import LabelEncoder
import statistics

warnings.filterwarnings("ignore")

#filename = 'F:\\protein_stability\\result\\TOTAL\\deeplearn_10fold_181_75_40.xlsx'
#D:\\senior4\\NewResearch\\PresentMethod\\Protein_stabilityshixiong\\checkpoint\\deeplearn_181_75_40_1_label_{i}_{k}.csv'
filename = 'LOCOVTP.xlsx'

def op_toexcelTrain(data, filename):
    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]
        ws.append(data)  # 每次写入一行
        wb.save(filename)

    else:
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb['Sheet']  # 创建子表
        ws.append(['PRED_label','TRUE_label'])
        ws.append(data)  # 每次写入一行
        wb.save(filename)
def op_toexcel(data, filename):
    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]
        ws.append(data)  
        wb.save(filename)

    else:
        wb = op.Workbook()  
        ws = wb['Sheet'] 
        ws.append(['MSE', 'MAE', 'RMSE', 'R2', 'PCC', 'P_value', 'Delta'])
        ws.append(data)
        wb.save(filename)


def data_generator(train_esm, train_prot,train_Energy, train_y, batch_size):
    L = train_esm.shape[0]

    while True:
        for i in range(0, L, batch_size):
            batch_esm = train_esm[i:i + batch_size].copy()
            batch_prot = train_prot[i:i + batch_size].copy()
            batch_Energy = train_Energy[i:i + batch_size].copy()
            batch_y = train_y[i:i + batch_size].copy()

            yield ([batch_esm, batch_prot,batch_Energy], batch_y)


def cross_validation(train_esm, train_1v, train_prot, train_label, test_esm, test_1v, test_prot, test_label, label):
    # 训练、验证each epoch的步长
    train_size = train_label.shape[0]
    batch_size = 16  # 16
    train_steps = train_size // batch_size

    print(f"\n第{label}次实验样本数量：Training samples: {train_esm.shape[0]}, Test samples: {test_esm.shape[0]}")
    print("测试样本数量及样本:", len(test_label))
    print(test_label)

    qa_model = get_model()
    valiBestModel = f'new_checkpoint221_loo_deeplearn_b16_474DSSS120_{label}.h5'

    checkpointer = tf.keras.callbacks.ModelCheckpoint(
        filepath=valiBestModel,
        monitor='loss',
        save_weights_only=True,
        verbose=1,
        save_best_only=True,
        mode='min',
    )

    early_stopping = tf.keras.callbacks.EarlyStopping(
        monitor='loss',
        patience=80,
        verbose=0,
        mode='min'
    )
    train_generator = data_generator(train_esm, train_1v, train_prot, train_label, batch_size)


    history_callback = qa_model.fit(
        train_generator,
        steps_per_epoch=train_steps,
        epochs=40,
        verbose=1,
        callbacks=[checkpointer, early_stopping],
        shuffle=True,
        workers=1
    )

    # 训练完成后进行测试
    test_pred = qa_model.predict([test_esm,test_1v,test_prot]).reshape(-1, )
    #evaluate_regression(test_pred, test_label)

    y_true_flat = np.ravel(test_label)
    y_pred_flat = np.ravel(test_pred)


    result_df = pd.DataFrame({'True Label': y_true_flat, 'Predicted Label': y_pred_flat})

    # 追加数据到统一的csv文件中
    result_df.to_csv(combined_result_file, mode='a', header=False, index=False)

    train_generator.close()
    del train_generator
    gc.collect()
    print(
        f"\n第{label}次实验结果：Loss: {history_callback.history['loss'][-1]:.4f}, RMSE: {history_callback.history['root_mean_squared_error'][-1]:.4f}")


if __name__ == "__main__":
    # 设置tensorflow来使用特定的GPU，并将动态分配GPU内存
    combined_result_file = 'LOCOVresult.csv'

    # 生成表头，只需要执行一次
    if not os.path.exists(combined_result_file):
        # 创建一个空的 DataFrame，只包含表头
        pd.DataFrame({
            'TRUE_label': [],
            'PRED_label': []
        }).to_csv(combined_result_file, index=False)
    os.environ["CUDA_VISIBLE_DEVICES"] = "0"
    #os.environ["CUDA_VISIBLE_DEVICES"] = "CPU"

    gpus = tf.config.experimental.list_physical_devices('GPU')
    tf.config.experimental.set_memory_growth(gpus[0], True)

    #all_esm = np.lib.format.open_memmap('../features_npy/RNA_total_181_esm.npy')
    #all_prot = np.lib.format.open_memmap('../features_npy/RNA_total_181_prot.npy')
    #all_label = np.lib.format.open_memmap('../features_npy/RNA_total_181_label.npy')
    all_esm = np.lib.format.open_memmap('./features_npy/S552/S552_181_esm.npy')
    all_1v = np.lib.format.open_memmap('./features_npy/S552/S552_181_esm1v.npy')
    all_prot = np.lib.format.open_memmap('./features_npy/S552/S552_181_prot.npy')
    all_label = np.lib.format.open_memmap('./features_npy/S552/S552_181_label.npy')
    all_PDb = np.lib.format.open_memmap('./features_npy/S552/S552_181_pdbid.npy')

    label_indices = defaultdict(list)
    for index, label in enumerate(all_PDb):
        label_indices[label].append(index)
    # 获取所有独特的类标
    unique_labels = list(label_indices.keys())

    print("交叉验证样本数量", len(all_label))
    all_label_train = all_label.astype(np.float)

    all_y = all_label_train
    num_samples = all_label.shape[0]
    all_label = all_label.astype(np.float) 

    print("all_esm shape:", all_esm.shape)
    print("all_1v shape:", all_1v.shape)
    print("all_prot shape:", all_prot.shape)
    print("all_label shape:", all_label.shape)

    for label in unique_labels:
        # 获取当前类标的测试集索引
        test_indices = np.array(label_indices[label])

        # 获取去除当前类标后的训练集索引
        train_indices = [i for i in range(len(all_PDb)) if all_PDb[i] != label]

        # 生成训练集和测试集
        train_ESM = all_esm[train_indices]
        train_Prot = all_prot[train_indices]
        train_Y = all_label[train_indices]
        train_1v =  all_1v[train_indices]

        test_ESM = all_esm[test_indices]
        test_Prot = all_prot[test_indices]
        test_Y = all_label[test_indices]
        test_1v = all_1v[test_indices]

        cross_validation(train_ESM, train_1v,train_Prot, train_Y, test_ESM, test_1v ,test_Prot, test_Y, label)